# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

#%% Modules
import numpy as np
import openpyxl
import pandas as pd
#%%Data
file = r"C:\Users\uazkg2\OneDrive - The University of Nottingham\Code\Chemistry_and_Physics_timesheet_list_Jan_25.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb['Sheet1']

#%% Getting column data
names = np.array([ws.cell(row=i, column=8).value for i in range(2, ws.max_row+1)])
names = np.delete(names, names=='ResID(T)')

# surnames = np.array([np.char.split(i, ',').tolist() for i in names])

surname, firstname = np.array([]), np.array([])

for i in names:
    name = np.array(np.char.split(i, ',').tolist())
    surname = np.append(surname, name[0])
    firstname = np.append(firstname, name[1])

#%% Copying to clipboard            
df = pd.DataFrame(firstname)
df.to_clipboard(index=False,header=False)