# -*- coding: utf-8 -*-
"""
Created on Tue Aug 12 09:31:22 2025

@author: uazkg2
"""

#%% Modules
from pathlib import Path
from pypdf import PdfReader, PdfWriter
import numpy as np
import openpyxl
import re
from datetime import datetime

#%% Data
pure_path = '/Users/uazkg2/OneDrive - The University of Nottingham/Documents/Taught_Projects/'
fp = Path(pure_path) #folder path
assert fp.exists()
files = np.array(list(fp.iterdir())) #List of file names

time = np.array([i.stat().st_ctime for i in files]) #Sorts files by time put in folder

target = files[np.where(time==np.max(time))][0] #Latest file put into the folder

#%% Read in the pdf file
supervisor = np.array([])
student = np.array([])

reader = PdfReader(target)
page = reader.pages[0] #normally the text i need will be on the first page
text = page.extract_text()
lines = text.splitlines()

for line in lines: #searching for supervisor name
    match = re.search(r"Name of Supervisor[:\-]?\s*(.+)", line, re.IGNORECASE)
    if match:        
        name = match.group(1).strip() #extracts the name from the match object
        supervisor = np.append(supervisor, name)

for line in lines: #searching for student name
    match = re.search(r"Names of the students conducting the research[:\-]?\s*(.+)", line, re.IGNORECASE)
    if match:        
        name = match.group(1).strip() #extracts the name from the match object
        student = np.append(student, name)

#%%Getting date for excel file
formatted_date = datetime.now().strftime("%d/%m/%Y")
       
#%% Excel file
workbook = openpyxl.load_workbook("/Users/uazkg2/OneDrive - The University of Nottingham/Psych-Ethics - Taught Project Ethical Review Forms/Ethics Archive Number - Taught Project Forms.xlsx")
worksheet = workbook.active

#%% Getting next empty cell
# d = worksheet.cell(row=87, column=3)
row=1
while True:
    cell_value = worksheet.cell(row=row, column=1).value
    if cell_value==None:
        break
    else:
        row+=1

#%% Getting archive number value
new_archive = worksheet.cell(row=(row-1), column=3).value+1

#%% Writing into the excel file

worksheet.cell(row=row, column=1, value=supervisor[0]+' / '+student[0])
worksheet.cell(row=row, column=2, value=formatted_date)
worksheet.cell(row=row, column=3, value=new_archive)

workbook.save('/Users/uazkg2/OneDrive - The University of Nottingham/Psych-Ethics - Taught Project Ethical Review Forms/Ethics Archive Number - Taught Project Forms.xlsx')

#%% Building the string for the name of the pdf
supervisor_surname = supervisor[0].split(' ')[1]
output_path = f"/Users/uazkg2/OneDrive - The University of Nottingham/Psych-Ethics - Taught Project Ethical Review Forms/Taught Project Review Forms 2024-2025/Taught Projects {new_archive} {supervisor_surname}.pdf"
#%% Saving PDF
writer = PdfWriter()

for page in reader.pages:
    writer.add_page(page)

with open(output_path, "wb") as output_file:
    writer.write(output_file)




















