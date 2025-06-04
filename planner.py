# -*- coding: utf-8 -*-
"""
Created on Tue Jun  3 09:15:46 2025
Readme: Code that edits the planner output.
@author: uazkg2
"""
#%% Modules
from pathlib import Path
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
import openpyxl.utils
from datetime import datetime
#%% functions
def row_values(row):
    row_data = next(ws.iter_rows(min_row=row, max_row=row, values_only=True))
    return np.array(row_data)

def col_values(col, title=True):
    if title==False:
        col_data = next(ws.iter_cols(min_col=col, max_col=col, values_only=True))
    else:
        col_data = next(ws.iter_cols(min_col=col, max_col=col, values_only=True))
        col_data = col_data[1:]
    return np.array(col_data)
#%% Data
folder_path = '/Users/uazkg2/OneDrive - The University of Nottingham/Documents/Planner'
fp = Path(folder_path) #folder path
assert fp.exists()

target = np.array(list(fp.iterdir()))[0]
assert target.exists()

wb = openpyxl.load_workbook(target)
ws = wb['May']
titles = row_values(1)

#%% Deadline - Splitting into a date and time
indx = int(np.where(titles=='Funder Deadline')[0][0]+1)
deadline = col_values(indx)

column = [get_column_letter(cell.column) for cell in ws[1] if cell.value == 'Funder Deadline']

dline = []

for i in deadline:
    if type(i)==str:
        dline.append(datetime.strptime(i, "%Y-%m-%d %H:%M"))
    else:
        dline.append(np.nan)

dline = np.stack(dline) #Turns list into numpy array

for i in zip(ws[column[0]][1:], dline):
    # print(i[0].value)
    if type(i[0].value) == type(None):
        continue
    else:
        i[0].value = i[1].strftime('%d-%m-%y')
    


    
#%%
wb.save(fp/'text.xlsx')

